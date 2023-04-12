from openpyxl import load_workbook
import openpyxl

file = 'test.xlsx'
wb = load_workbook(file)
ws = wb.active

cols = [
        ['Дет.', '', 2, 'шт.', '', 'Сталь 10  ГОСТ 1050-2013 2.0 мм'],
        ['Дет.', '', 2, 'шт.', '', 'Сталь 10  ГОСТ 1050-2013 2.0 мм'], 
        ['Ст.изд.', '', 1, 'шт.', '', 'Без указания материала'], 
        ['Ст.изд.', '', 1, 'шт.', '', 'Без указания материала'],
        ['Сб.ед.', 'низ с профилем', 1, 'шт.', 'Сборка платформа упрощенная', ''], 
        ['Сб.ед.', 'низ с профилем', 1, 'шт.', 'Сборка платформа упрощенная', '']
        ]

for row_txt in range(7, 7+len(cols)):
      txt=cols[row_txt-7]
      a=0
      for col_txt in [3,4,5,6,8,9]:
          value_txt = txt[a]
          a=a+1
          ws.cell(row=row_txt, column=col_txt).value = value_txt

wb.save(file)