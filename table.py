from openpyxl import load_workbook, Workbook
import openpyxl
import os

# Функция открытия и записи в таблицу значений, где path - путь до таблицы, cols - массив с массивами строк значений
def save_file(path, settings, info, entry):     
    if path == '':
        wb = Workbook()
        path = 'new_specification.xlsx'
    else:
        wb = load_workbook(path)
    
    row = 7
    columns = []
    if settings[0]: columns.append(2)
    if settings[1]: columns.append(3)
    if settings[2]: columns.append(4)
    if settings[3]: 
        columns.append(5)
        columns.append(6)
    if settings[4]: columns.append(8)
    if settings[5]: columns.append(9)
    if settings[6]: columns.append(10)
    if settings[7]: columns.append(11)

    titles = []

    workSheet = wb.active  # Получим активный лист таблицы
    try:
        workSheet.title = info[0][0][5]
    except:
        workSheet.title = 'Лист №1'
    titles.append(workSheet.title)

    for i in range(1, len(info)):
        copySheet = wb.copy_worksheet(workSheet)
        try:
            copySheet.title = info[i][0][5]
        except:
            copySheet.title = 'Лист №' + str(i+1)
        titles.append(copySheet.title)    
    
    for i in range(len(info)):
        wb.active = wb[titles[i]]
        workSheet = wb.active
        paste_info(workSheet, info[i], row, columns)
        workSheet.cell(row=5, column=1).value = entry
    
    wb.save(path)
    os.startfile(path)

def paste_info(workSheet, info, row, columns):
    for row_txt in range(row, row + len(info)): 
        txt=info[row_txt-row]
        a=0
        for col_txt in columns:
            value_txt = txt[a]
            a=a+1
            workSheet.cell(row=row_txt, column=col_txt).value = value_txt